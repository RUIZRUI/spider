import openpyxl

readWb = openpyxl.load_workbook('面向对象方法与C程序-07-平时练习统计详情.xlsx')
writeWb = openpyxl.load_workbook('2022年C++平时成绩.xlsx')
excelName = '2022年C++平时成绩.xlsx'

## 定义变量
# 学号列表
sidList = []
# 分数列表
scoreList = []
# sheet数量
sheetNumber = 9
# 学生数量
stuNumber = 60
# 每个sheet总分
totalScoreList = [10, 10, 10, 20, 5, 10, 10, 20, 2]
# 分数列索引（从0开始）
columnIndex = 6



def readSid():
    '''
    读取《2022年C++平时成绩.xlsx》的学号列表
    '''
    sheet = writeWb.worksheets[0]

    # 遍历
    for row in sheet.iter_rows():
        counter = 0
        for cell in row:
            if counter == 0 and (cell.value == None or cell.value == '学号' or cell.value == ''):
                # 无效数据，跳出循环
                break
            if counter == 0:
                sidList.append(str(cell.value))
                counter += 1
                break
    # print(sidList)
    print(len(sidList))


def readScore():
    '''
    读取《面向对象方法与C程序-07-平时练习统计详情.xlsx》的分数列表
    '''
    # 遍历 sheet
    for index in range(sheetNumber):
        sheet = readWb.worksheets[index]
        # 遍历分数
        rowCounter = 0
        sList = []
        for row in sheet.iter_rows():
            counter = 0
            for cell in row:
                if rowCounter < 6:
                    # 无效数据，跳出循环
                    break
                if counter == 0:
                    # print('row=', rowCounter, ', value=', cell.value, end=' ')
                    if sidList[rowCounter-6] == str(cell.value):
                        # print(', 正确', end=' ')
                        pass
                    else:
                        # print(', 错误', end=' ')
                        pass
                if counter == columnIndex:
                    if cell.value == None or cell.value == '':
                        # print('None')
                        sList.append(0)
                        # print(', score=', str(0))
                    else:
                        sList.append(int(cell.value))
                        # print(', score=', str(cell.value))
                counter += 1
            rowCounter += 1
        # print(sList)
        print(len(sList))
        scoreList.append(sList)
    # print(scoreList)
    print(len(scoreList))


def sortByScore():
    '''
    通过分数进行排序
    '''
    for i in range(stuNumber):
        originalList = []
        for j in range(sheetNumber):
            originalList.append(scoreList[j][i])
        print(originalList)

        # 复制 originalList
        sList = []
        for j in range(len(originalList)):
            sList.append((originalList[j]/totalScoreList[j])*4)
        print(sList)
        # 排序
        quickSort(originalList, sList, 0, len(sList)-1)
        print(originalList)
        print(sList)
        # 将排序后的分数写入 Excel 文件 《2022年C++平时成绩.xlsx》
        writeScore(originalList, sList, i)


def writeScore(originalList, sList, stuIndex):
    '''
    将排序后的分数写入 Excel 文件 《2022年C++平时成绩.xlsx》
    '''
    length = len(sList)
    # 挑选最高的5个分数
    sum = 0
    for i in range(5):
        print(originalList[length-i-1], end=',')
        sum += sList[length-i-1]
    print(sum)

    sheet = writeWb.worksheets[0]
    # 遍历
    rowIndex = 0
    for row in sheet.iter_rows():
        counter = 0
        if rowIndex != stuIndex+2:
            # 写入指定行
            # print('rowIndex=', rowIndex, ', stuIndex=', stuIndex)
            rowIndex += 1
            continue
        print('rowIndex=', rowIndex, ', stuIndex=', stuIndex)
        for cell in row:
            if counter == 0 and (cell.value == None or cell.value == '学号'):
                # 无效数据，跳出循环
                break
            if counter == 20:
                # 平时1
                print(originalList[length - 1], end=' ')
                cell.value = originalList[length-1]
            elif counter == 21:
                # 平时2
                print(originalList[length - 2], end=' ')
                cell.value = originalList[length - 2]
            elif counter == 22:
                # 平时3
                print(originalList[length - 3], end=' ')
                cell.value = originalList[length - 3]
            elif counter == 23:
                # 平时4
                print(originalList[length - 4], end=' ')
                cell.value = originalList[length - 4]
            elif counter == 24:
                # 平时5
                print(originalList[length - 5], end=' ')
                cell.value = originalList[length - 5]
            elif counter == 25:
                # 小分
                print(sum)
                cell.value = sum
            counter += 1
        rowIndex += 1

    # 保存
    writeWb.save(excelName)









def quickSort(originalList, sList, left, right):
    '''
    快排
    '''
    if left < right:
        position = partition(originalList, sList, left, right)
        quickSort(originalList, sList, left, position - 1)
        quickSort(originalList, sList, position + 1, right)


def partition(originalList, sList, left, right):
    '''
    快排 partition
    '''
    i = left
    j = left
    while j <= right:
        if sList[j] >= sList[left]:
            j += 1
        else:
            i += 1
            temp = sList[i]
            sList[i] = sList[j]
            sList[j] = temp
            temp = originalList[i]
            originalList[i] = originalList[j]
            originalList[j] = temp
            j += 1
    temp = sList[left]
    sList[left] = sList[i]
    sList[i] = temp
    temp = originalList[left]
    originalList[left] = originalList[i]
    originalList[i] = temp
    return i





if __name__ == '__main__':
    readSid()
    readScore()
    sortByScore()





